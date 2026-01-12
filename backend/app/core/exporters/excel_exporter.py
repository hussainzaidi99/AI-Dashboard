"""
Excel Exporter - Export data and charts to Excel format
Creates Excel workbooks with data, charts, and formatting
"""
#backend/app/core/exporters/excel_exporter.py


from typing import List, Optional, Dict, Any, Union
import os
from datetime import datetime
import pandas as pd
import plotly.graph_objects as go
import logging

from app.config import settings
from .base_exporter import BaseExporter

logger = logging.getLogger(__name__)


class ExcelExporter(BaseExporter):
    """Export data and charts to Excel format"""
    
    def __init__(self):
        super().__init__()
        
        # Check if feature is enabled
        self._ensure_enabled(settings.ENABLE_EXPORT_EXCEL, "Excel")
    
    def export_dataframe(
        self,
        df: pd.DataFrame,
        output_path: str,
        sheet_name: str = "Data",
        **kwargs
    ) -> str:
        """
        Export single DataFrame to Excel
        
        Args:
            df: pandas DataFrame
            output_path: Output file path
            sheet_name: Sheet name
            **kwargs: Additional pandas to_excel options
        
        Returns:
            Path to exported file
        """
        try:
            # Ensure output directory exists
            self._ensure_output_dir(output_path)
            
            # Export to Excel
            df.to_excel(
                output_path,
                sheet_name=sheet_name,
                index=kwargs.get('index', False),
                **{k: v for k, v in kwargs.items() if k != 'index'}
            )
            
            self.logger.info(f"Exported DataFrame to Excel: {output_path}")
            return output_path
        
        except Exception as e:
            self.logger.error(f"Error exporting DataFrame to Excel: {str(e)}")
            raise
    
    def export_multiple(
        self,
        dataframes: Dict[str, pd.DataFrame],
        output_path: str,
        **kwargs
    ) -> str:
        """
        Export multiple DataFrames to Excel (each as separate sheet)
        
        Args:
            dataframes: Dictionary of {sheet_name: DataFrame}
            output_path: Output file path
            **kwargs: Additional options
        
        Returns:
            Path to exported file
        """
        try:
            # Ensure output directory exists
            self._ensure_output_dir(output_path)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(
                        writer,
                        sheet_name=sheet_name[:31],  # Excel sheet name limit
                        index=kwargs.get('index', False)
                    )
            
            self.logger.info(f"Exported {len(dataframes)} sheets to Excel: {output_path}")
            return output_path
        
        except Exception as e:
            self.logger.error(f"Error exporting multiple DataFrames: {str(e)}")
            raise
    
    def export_with_charts(
        self,
        dataframes: Union[List[pd.DataFrame], Dict[str, pd.DataFrame]],
        figures: List[go.Figure],
        output_path: str,
        **kwargs
    ) -> str:
        """
        Export DataFrames with embedded chart images
        
        Args:
            dataframes: List or dict of DataFrames
            figures: List of Plotly Figure objects
            output_path: Output file path
            **kwargs: Additional options
        
        Returns:
            Path to exported file
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XLImage
            
            # Ensure output directory exists
            self._ensure_output_dir(output_path)
            
            # Convert list to dict if necessary
            if isinstance(dataframes, list):
                dataframes = {f"Data_{i+1}": df for i, df in enumerate(dataframes)}
            
            # Create Excel file with data
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(
                        writer,
                        sheet_name=sheet_name[:31],
                        index=False
                    )
            
            # Load workbook to add images
            wb = load_workbook(output_path)
            
            # Use temp directory context manager for images
            with self._temp_dir() as temp_dir:
                # Add charts as images
                for idx, fig in enumerate(figures):
                    # Create or get sheet
                    if idx < len(dataframes):
                        sheet_name = list(dataframes.keys())[idx][:31]
                        ws = wb[sheet_name]
                    else:
                        # Create new sheet for extra charts
                        sheet_name = f"Chart_{idx+1}"
                        ws = wb.create_sheet(sheet_name)
                    
                    # Export figure to temp image
                    temp_image_path = os.path.join(temp_dir, f"chart_{idx}.png")
                    fig.write_image(
                        temp_image_path,
                        format="png",
                        width=800,
                        height=500,
                        scale=2,
                        engine="kaleido"
                    )
                    
                    # Add image to Excel
                    img = XLImage(temp_image_path)
                    
                    # Position image below data (if data exists)
                    if idx < len(dataframes):
                        df = list(dataframes.values())[idx]
                        start_row = len(df) + 5  # Leave some space
                    else:
                        start_row = 2
                    
                    ws.add_image(img, f'A{start_row}')
                
                # Save workbook
                wb.save(output_path)
            
            self._log_export_success(output_path, "Excel with charts")
            return output_path
        
        except ImportError:
            raise RuntimeError(
                "Excel export with charts requires openpyxl and kaleido. "
                "Install with: pip install openpyxl kaleido"
            )
        except Exception as e:
            self._log_export_error(e, "Excel with charts")
            raise
    
    def export_formatted(
        self,
        df: pd.DataFrame,
        output_path: str,
        sheet_name: str = "Data",
        auto_filter: bool = True,
        freeze_header: bool = True,
        column_width: Optional[int] = None,
        **kwargs
    ) -> str:
        """
        Export DataFrame with formatting
        
        Args:
            df: pandas DataFrame
            output_path: Output file path
            sheet_name: Sheet name
            auto_filter: Enable auto-filter
            freeze_header: Freeze header row
            column_width: Column width (auto-size if None)
            **kwargs: Additional options
        
        Returns:
            Path to exported file
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Ensure output directory exists
            self._ensure_output_dir(output_path)
            
            # Export to Excel
            df.to_excel(output_path, sheet_name=sheet_name, index=False)
            
            # Load workbook for formatting
            wb = load_workbook(output_path)
            ws = wb[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-filter
            if auto_filter:
                ws.auto_filter.ref = ws.dimensions
            
            # Freeze header
            if freeze_header:
                ws.freeze_panes = "A2"
            
            # Column width
            if column_width:
                for column in ws.columns:
                    ws.column_dimensions[column[0].column_letter].width = column_width
            else:
                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                   min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            # Alternating row colors
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
                if idx % 2 == 0:
                    fill = PatternFill(
                        start_color="F2F2F2",
                        end_color="F2F2F2",
                        fill_type="solid"
                    )
                    for cell in row:
                        cell.fill = fill
            
            # Save workbook
            wb.save(output_path)
            
            self.logger.info(f"Exported formatted Excel: {output_path}")
            return output_path
        
        except ImportError:
            raise RuntimeError(
                "Formatted Excel export requires openpyxl. "
                "Install with: pip install openpyxl"
            )
        except Exception as e:
            self.logger.error(f"Error exporting formatted Excel: {str(e)}")
            raise
    
    def export_with_summary(
        self,
        df: pd.DataFrame,
        output_path: str,
        summary_stats: bool = True,
        charts: Optional[List[go.Figure]] = None,
        **kwargs
    ) -> str:
        """
        Export DataFrame with summary sheet
        
        Args:
            df: pandas DataFrame
            output_path: Output file path
            summary_stats: Include summary statistics
            charts: Optional list of charts to include
            **kwargs: Additional options
        
        Returns:
            Path to exported file
        """
        try:
            # Create summary
            summary_data = {
                'Metric': ['Total Rows', 'Total Columns', 'Export Date'],
                'Value': [
                    len(df),
                    len(df.columns),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            
            # Add statistical summary if requested
            if summary_stats:
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    stats_df = df[numeric_cols].describe().T
                    stats_df = stats_df.reset_index()
                    stats_df.rename(columns={'index': 'Column'}, inplace=True)
            
            # Create Excel file
            dataframes = {
                'Summary': summary_df,
                'Data': df
            }
            
            if summary_stats and len(numeric_cols) > 0:
                dataframes['Statistics'] = stats_df
            
            # Export with or without charts
            if charts:
                return self.export_with_charts(dataframes, charts, output_path, **kwargs)
            else:
                return self.export_multiple(dataframes, output_path, **kwargs)
        
        except Exception as e:
            self.logger.error(f"Error exporting Excel with summary: {str(e)}")
            raise
    
    def append_to_existing(
        self,
        df: pd.DataFrame,
        file_path: str,
        sheet_name: str,
        **kwargs
    ) -> str:
        """
        Append DataFrame to existing Excel file
        
        Args:
            df: pandas DataFrame
            file_path: Existing Excel file path
            sheet_name: Sheet name to append to (or create)
            **kwargs: Additional options
        
        Returns:
            Path to Excel file
        """
        try:
            from openpyxl import load_workbook
            
            if not os.path.exists(file_path):
                # Create new file if doesn't exist
                return self.export_dataframe(df, file_path, sheet_name, **kwargs)
            
            # Load existing workbook
            with pd.ExcelWriter(
                file_path,
                engine='openpyxl',
                mode='a',
                if_sheet_exists='overlay'
            ) as writer:
                # Get existing data if sheet exists
                try:
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                    # Find next empty row
                    startrow = len(existing_df) + 1
                except:
                    startrow = 0
                
                # Append data
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=startrow,
                    index=False,
                    header=(startrow == 0)
                )
            
            self.logger.info(f"Appended data to Excel: {file_path}")
            return file_path
        
        except Exception as e:
            self.logger.error(f"Error appending to Excel: {str(e)}")
            raise