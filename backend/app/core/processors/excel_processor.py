"""
Excel Processor - Extract data from Excel files (.xlsx, .xls)
Supports multiple sheets, formulas, and formatting
"""
#backend/app/core/processors/excel_processor.py


import time
from typing import List, Optional, Dict, Any
import pandas as pd
import numpy as np
import logging

from .base import BaseProcessor, ProcessingResult
from app.config import settings

logger = logging.getLogger(__name__)


class ExcelProcessor(BaseProcessor):
    """Process Excel files and extract data from all sheets"""
    
    def __init__(self):
        super().__init__()
    
    def process(self, file_path: str, **kwargs) -> ProcessingResult:
        """
        Process Excel file and extract all sheets
        
        Args:
            file_path: Path to Excel file
            **kwargs:
                - sheet_name: str or int or list (default: None = all sheets)
                - header: int or None (default: 0)
                - skiprows: int (default: 0)
                - usecols: str or list (columns to use)
        
        Returns:
            ProcessingResult with extracted data
        """
        start_time = time.time()
        
        try:
            # Validate file
            self.validate_file(file_path)
            
            # Get file metadata
            metadata = self.get_file_metadata(file_path)
            
            # Extract options
            sheet_name = kwargs.get('sheet_name')
            header = kwargs.get('header', 0)
            skiprows = kwargs.get('skiprows', 0)
            usecols = kwargs.get('usecols')
            
            # Determine Excel engine based on file extension
            extension = file_path.lower().split('.')[-1]
            engine = 'openpyxl' if extension in ['xlsx', 'xlsm', 'xltx', 'xltm'] else 'xlrd'
            
            # Auto-detect header if default (0) and not explicitly provided in kwargs
            if 'header' not in kwargs:
                # Detect header for each sheet if sheet_name is None (all sheets)
                # or for the specific sheet
                try:
                    detected_header = self._detect_header_row(file_path, sheet_name, engine=engine)
                    if detected_header is not None:
                        header = detected_header
                        self.logger.info(f"Auto-detected header at row {header} for {file_path}")
                except Exception as he:
                    self.logger.warning(f"Header detection failed: {str(he)}")

            # Read Excel file
            dataframes = []
            sheet_names = []
            warnings = []
            
            if engine == 'xlrd' and not file_path.endswith('.xls'):
                warnings.append(f"Using xlrd engine for unknown extension. This may fail on modern Excel files.")
            
            try:
                # Read all sheets or specific sheet
                excel_data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=header,
                    skiprows=skiprows,
                    usecols=usecols,
                    engine=engine
                )
                
                # Handle single sheet vs multiple sheets
                if isinstance(excel_data, dict):
                    # Multiple sheets
                    for sheet, df in excel_data.items():
                        if not df.empty:
                            cleaned_df = self.clean_dataframe(df)
                            dataframes.append(cleaned_df)
                            sheet_names.append(str(sheet))
                else:
                    # Single sheet
                    if not excel_data.empty:
                        cleaned_df = self.clean_dataframe(excel_data)
                        dataframes.append(cleaned_df)
                        sheet_names.append(
                            str(sheet_name) if sheet_name else "Sheet1"
                        )
                
                # Get additional metadata using openpyxl
                if file_path.endswith('.xlsx'):
                    excel_metadata = self._extract_excel_metadata(file_path)
                    metadata.update(excel_metadata)
            
            except Exception as e:
                import traceback
                error_detail = traceback.format_exc().splitlines()[-1]
                self.logger.error(f"Error reading Excel file: {str(e)}")
                warnings.append(f"Error reading sheets: {error_detail}")
                
                if engine == 'xlrd' and "Excel 2007" in str(e):
                    warnings.append("Note: xlrd only supports old .xls files. For .xlsx, openpyxl is required.")
            
            if not dataframes:
                raise ValueError("No data could be extracted from Excel file")
            
            # Create result
            result = ProcessingResult(
                success=True,
                file_path=file_path,
                file_type='excel',
                dataframes=dataframes,
                metadata=metadata,
                processing_time=time.time() - start_time,
                warnings=warnings,
                sheet_names=sheet_names
            )
            
            self.log_processing_info(result)
            return result
        
        except Exception as e:
            self.logger.error(f"Error processing Excel {file_path}: {str(e)}")
            return ProcessingResult(
                success=False,
                file_path=file_path,
                file_type='excel',
                error_message=str(e),
                processing_time=time.time() - start_time
            )
    
    def _extract_excel_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        Extract metadata from Excel file using openpyxl
        
        Args:
            file_path: Path to Excel file
        
        Returns:
            Dictionary with metadata
        """
        metadata = {}
        
        try:
            from openpyxl import load_workbook
            
            # Load workbook (read-only for performance)
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            # Get sheet information
            metadata['sheet_count'] = len(wb.sheetnames)
            metadata['sheet_names'] = wb.sheetnames
            
            # Get workbook properties
            props = wb.properties
            if props:
                metadata['created'] = str(props.created) if props.created else None
                metadata['modified'] = str(props.modified) if props.modified else None
                metadata['creator'] = props.creator
                metadata['last_modified_by'] = props.lastModifiedBy
                metadata['title'] = props.title
                metadata['subject'] = props.subject
            
            wb.close()
        
        except Exception as e:
            self.logger.warning(f"Could not extract Excel metadata: {str(e)}")
        
        return metadata
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get all sheet names from Excel file
        
        Args:
            file_path: Path to Excel file
        
        Returns:
            List of sheet names
        """
        try:
            engine = 'openpyxl' if file_path.endswith('.xlsx') else 'xlrd'
            excel_file = pd.ExcelFile(file_path, engine=engine)
            return excel_file.sheet_names
        except Exception as e:
            self.logger.error(f"Error getting sheet names: {str(e)}")
            return []
    
    def process_specific_sheet(
        self,
        file_path: str,
        sheet_name: str
    ) -> ProcessingResult:
        """
        Process only a specific sheet from Excel file
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of the sheet to process
        
        Returns:
            ProcessingResult with single sheet data
        """
        return self.process(file_path, sheet_name=sheet_name)
    
    def process_with_formulas(self, file_path: str) -> Dict[str, Any]:
        """
        Process Excel file and extract formulas (not just values)
        
        Args:
            file_path: Path to Excel file
        
        Returns:
            Dictionary with sheet data including formulas
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, data_only=False)
            
            sheets_data = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                sheet_info = {
                    'data': [],
                    'formulas': {},
                    'dimensions': ws.dimensions
                }
                
                # Extract data and formulas
                for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
                    row_data = []
                    for col_idx, cell in enumerate(row, 1):
                        row_data.append(cell.value)
                        
                        # Check if cell has a formula
                        if cell.data_type == 'f':
                            cell_coord = f"{chr(64 + col_idx)}{row_idx}"
                            sheet_info['formulas'][cell_coord] = cell.value
                    
                    sheet_info['data'].append(row_data)
                
                sheets_data[sheet_name] = sheet_info
            
            wb.close()
            return sheets_data
        
        except Exception as e:
            self.logger.error(f"Error extracting formulas: {str(e)}")
            return {}
    
    def detect_data_range(self, file_path: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        Detect the actual data range in Excel sheet (skip empty rows/cols)
        
        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name (optional)
        
        Returns:
            Dictionary with data range information
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active
            
            # Get dimensions
            min_row = sheet.min_row
            max_row = sheet.max_row
            min_col = sheet.min_column
            max_col = sheet.max_column
            
            # Find actual data range (skip empty leading rows/cols)
            actual_min_row = min_row
            actual_min_col = min_col
            
            # Find actual min row
            for row_idx in range(min_row, max_row + 1):
                row_cells = [sheet.cell(row=row_idx, column=c).value for c in range(min_col, max_col + 1)]
                if any(v is not None for v in row_cells):
                    actual_min_row = row_idx
                    break
            
            # Find actual min column
            for col_idx in range(min_col, max_col + 1):
                col_cells = [sheet.cell(row=r, column=col_idx).value for r in range(min_row, max_row + 1)]
                if any(v is not None for v in col_cells):
                    actual_min_col = col_idx
                    break
            
            wb.close()
            
            return {
                'min_row': actual_min_row,
                'max_row': max_row,
                'min_col': actual_min_col,
                'max_col': max_col,
                'total_rows': max_row - actual_min_row + 1,
                'total_cols': max_col - actual_min_col + 1
            }
        
        except Exception as e:
            self.logger.error(f"Error detecting data range: {str(e)}")
            return {}

    def _detect_header_row(self, file_path: str, sheet_name: Any = None, engine: str = 'openpyxl') -> Optional[int]:
        """
        Heuristic to detect the most likely header row in an Excel sheet.
        """
        try:
            # Read first 20 rows without header to analyze
            preview_df = pd.read_excel(
                file_path,
                sheet_name=sheet_name if sheet_name is not None else 0,
                header=None,
                nrows=20,
                engine=engine
            )
            
            if preview_df.empty:
                return None
            
            scores = []
            for idx, row in preview_df.iterrows():
                non_null_row = row.dropna()
                # Also treat empty strings as null for heuristic purposes
                non_empty_row = non_null_row[non_null_row.astype(str).str.strip() != ""]
                
                if non_empty_row.empty:
                    scores.append(-1.0)
                    continue
                
                # Factors for a good header:
                # 1. Breadth: Headers usually span multiple columns
                width_ratio = non_empty_row.count() / len(row)
                
                # 2. Type: Most values are strings (headers are labels)
                string_ratio = non_empty_row.apply(lambda x: isinstance(x, str)).mean()
                
                # 3. Quality: Values are unique
                unique_ratio = non_empty_row.nunique() / non_empty_row.count() if non_empty_row.count() > 0 else 0
                
                # 4. Length: Values aren't too long
                avg_len = non_empty_row.apply(lambda x: len(str(x))).mean()
                len_score = 1.0 if avg_len < 30 else (0.5 if avg_len < 60 else 0.1)
                
                # 5. Penalize numeric-heavy rows
                numeric_count = pd.to_numeric(non_empty_row, errors='coerce').notna().sum()
                numeric_ratio = numeric_count / non_empty_row.count()
                
                # Calculate weighted score (Heuristic)
                # Stronger weights on width and strings
                score = (width_ratio * 0.4) + (string_ratio * 0.3) + (unique_ratio * 0.2) + (len_score * 0.1) - (numeric_ratio * 0.6)
                scores.append(score)
            
            if not scores:
                return None
                
            # Pick the best overall row
            best_row = int(np.argmax(scores))
            best_score = scores[best_row]
            
            # Threshold Check
            if best_score > 0.4:
                # If there's a tie or close calls, the earlier one is usually a title, 
                # but our width ratio should have penalized it.
                # However, if two rows look like headers, the one with more content is likely the real header.
                return best_row
                
            return None
        except Exception as e:
            self.logger.debug(f"Header detection internal error: {str(e)}")
            return None
